# -*- coding: utf-8 -*-
"""WAVE Lounge Chair — Engineering BOM (concept / EBOM)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle, Protection,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import DataPoint as DP

OUT = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Palette (match WAVE-LC-001 drawing)
# ---------------------------------------------------------------------------
C_INK = "1B1B1B"
C_HEAD = "2C241C"
C_HEAD2 = "4A3B2A"
C_ACCENT = "8B5A2B"
C_OAK = "E4C9A0"
C_CREAM = "F3EFE6"
C_PAPER = "FBF9F4"
C_WHITE = "FFFFFF"
C_LINE = "C8C2B6"
C_GREEN = "2E6B4F"
C_MUTED = "6B6560"
C_ROW_ALT = "F7F1E8"
C_CAT = {
    "結構木材": "E8D4B5",
    "接合件": "D9E2D4",
    "五金": "D4DCE6",
    "膠合": "E6DCC8",
    "塗裝": "E8D9C8",
    "軟包": "F0EBE3",
    "包裝": "E4E4E0",
    "工裝攤提": "EDE4D8",
}

THIN = Border(
    left=Side(style="thin", color=C_LINE),
    right=Side(style="thin", color=C_LINE),
    top=Side(style="thin", color=C_LINE),
    bottom=Side(style="thin", color=C_LINE),
)
MED = Border(
    left=Side(style="medium", color=C_HEAD),
    right=Side(style="medium", color=C_HEAD),
    top=Side(style="medium", color=C_HEAD),
    bottom=Side(style="medium", color=C_HEAD),
)

FONT_NAME = "Microsoft JhengHei"


def font(size=10, bold=False, color=C_INK, name=FONT_NAME):
    return Font(name=name, size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# BOM data — 1 finished chair (WAVE-LC-001 Rev.A)
# qty = per chair; cost = mid estimate TWD for small batch (10 pcs)
# cost_lo / cost_hi = range
# ---------------------------------------------------------------------------
BOM = [
    # 結構木材
    dict(pn="WAVE-A01", name="左殼（扶手＋側腹＋左前舌＋左後踵）", cat="結構木材",
         spec="歐洲／美國白橡 層積實木，單層 22–28 mm，順紋沿緞帶",
         blank="毛坯 500 × 900 × 220", qty=1, unit="件",
         src="自製", lo=4800, mid=6200, hi=8200,
         note="含水率 8–10%；5 軸 OP1–OP3；最薄緣 ≥22 mm"),
    dict(pn="WAVE-B01", name="右殼（與 A01 對稱）", cat="結構木材",
         spec="同 A01，成對選料、色號一致",
         blank="毛坯 500 × 900 × 220", qty=1, unit="件",
         src="自製", lo=4800, mid=6200, hi=8200,
         note="與 A01 同批出材，避免左右色差"),
    dict(pn="WAVE-C01", name="座盆／脊板", cat="結構木材",
         spec="白橡層積，盆底等效厚 ≥32 mm",
         blank="毛坯 560 × 620 × 80", qty=1, unit="件",
         src="自製", lo=1100, mid=1500, hi=2100,
         note="止口 8–10 mm 與 A/B 對接；軟包完全遮縫"),
    dict(pn="WAVE-W09", name="色差備料／補片（同批）", cat="結構木材",
         spec="白橡，與主材同批、同含水",
         blank="約 0.020 m³", qty=1, unit="批",
         src="自製", lo=700, mid=1000, hi=1400,
         note="薄緣崩缺、端紋補片用；小量備料係數 1.15"),

    # 接合件
    dict(pn="WAVE-J01", name="圓榫／鬆榫", cat="接合件",
         spec="山毛櫸 Ø8 × 40，或 Domino 8×50（結構節點改 14×75）",
         blank="—", qty=18, unit="支",
         src="外購", lo=80, mid=150, hi=280,
         note="間距 80–110 mm；銷孔 H7，位置 ±0.08"),
    dict(pn="WAVE-J02", name="對位銷（組裝治具用，成品不留）", cat="接合件",
         spec="鋼銷 Ø8 h6，或工藝孔內木銷後切掉",
         blank="—", qty=4, unit="支",
         src="外購", lo=20, mid=40, hi=60,
         note="僅組裝；成品不可外露"),

    # 五金
    dict(pn="WAVE-H01", name="隱藏調整腳墊", cat="五金",
         spec="Ø20–25 mm，M8 牙，調程 ±3 mm，底附氈",
         blank="—", qty=4, unit="顆",
         src="外購", lo=80, mid=120, hi=200,
         note="消化接地平面度 0.5 mm／全長"),
    dict(pn="WAVE-H02", name="防刮氈墊（備用）", cat="五金",
         spec="Ø25 自黏羊毛氈，厚 3 mm",
         blank="—", qty=4, unit="片",
         src="外購", lo=10, mid=20, hi=40,
         note="若 H01 已附氈可省略"),
    dict(pn="WAVE-H03", name="軟包定位（魔鬼氈或暗銷）", cat="五金",
         spec="3M Dual Lock 或 Ø6 木暗銷 4 點",
         blank="—", qty=1, unit="套",
         src="外購", lo=30, mid=60, hi=120,
         note="坐墊可拆洗；不可用外露螺絲"),

    # 膠合
    dict(pn="WAVE-G01", name="層積膠", cat="膠合",
         spec="EPI 或 PVAc D4，實木層積",
         blank="約 0.80 kg／椅", qty=0.80, unit="kg",
         src="外購", lo=140, mid=200, hi=280,
         note="層間交錯年輪；加壓養護 ≥4 h"),
    dict(pn="WAVE-G02", name="結構組裝膠", cat="膠合",
         spec="室溫環氧或高固形 PU（端紋／止口）",
         blank="約 80 g／椅", qty=0.08, unit="kg",
         src="外購", lo=80, mid=140, hi=220,
         note="白膠不建議用於端紋對接"),
    dict(pn="WAVE-G03", name="填孔劑（橡木導管）", cat="膠合",
         spec="油性／水性木紋填孔，淺橡色",
         blank="約 40 g／椅", qty=0.04, unit="kg",
         src="外購", lo=30, mid=50, hi=80,
         note="開放漆或硬質油前使用"),

    # 塗裝
    dict(pn="WAVE-F01", name="硬質油／硬蠟油", cat="塗裝",
         spec="Osmo / Rubio 類，原色或白橡透明",
         blank="約 150–180 ml／椅（2–3 遍）", qty=0.17, unit="L",
         src="外購", lo=180, mid=280, hi=420,
         note="薄緣避免厚 PU；端紋先封"),
    dict(pn="WAVE-F02", name="砂紙耗材", cat="塗裝",
         spec="80 / 120 / 180 / 240 / 320，布基＋海綿砂",
         blank="約 8–12 張等效", qty=1, unit="套",
         src="外購", lo=70, mid=110, hi=180,
         note="CNC 留砂光餘量 0.25–0.35 mm"),
    dict(pn="WAVE-F03", name="拋光布／吸塵耗材", cat="塗裝",
         spec="無紡布、黏塵布",
         blank="—", qty=1, unit="套",
         src="外購", lo=20, mid=35, hi=60,
         note="上油前必須無粉塵"),

    # 軟包
    dict(pn="WAVE-U01", name="成型坐墊泡棉", cat="軟包",
         spec="冷熟成型 HR，55–70 kg/m³，厚 60–90 mm",
         blank="依座盆陰模", qty=1, unit="件",
         src="外購", lo=600, mid=1100, hi=1800,
         note="打樣可 CNC 切割泡棉；小量建議開模"),
    dict(pn="WAVE-U02", name="靠背成型泡棉", cat="軟包",
         spec="同 U01，可與坐墊一體或分件",
         blank="依靠背內曲面", qty=1, unit="件",
         src="外購", lo=500, mid=900, hi=1500,
         note="腰椎曲線做在泡棉，不做在木殼"),
    dict(pn="WAVE-U03", name="表布／皮革", cat="軟包",
         spec="頭層牛皮約 1.0 m²，或超纖 1.3 m",
         blank="含損耗", qty=1, unit="套",
         src="外購", lo=800, mid=2800, hi=5200,
         note="低配超纖；中高配頭層；色號與木殼對樣"),
    dict(pn="WAVE-U04", name="底布＋襯材", cat="軟包",
         spec="無紡布、定型棉 120–200 g、暗拉鍊 #5",
         blank="—", qty=1, unit="套",
         src="外購", lo=80, mid=150, hi=260,
         note="外輪廓退木唇 8–12 mm"),

    # 包裝
    dict(pn="WAVE-P01", name="外箱", cat="包裝",
         spec="BC 楞，內徑約 1020 × 1080 × 900（整椅出貨）",
         blank="—", qty=1, unit="只",
         src="外購", lo=160, mid=230, hi=320,
         note="若 A/B 拆裝出貨可改 2 箱縮小"),
    dict(pn="WAVE-P02", name="EPE／護角", cat="包裝",
         spec="EPE 20–30 mm ＋ 紙護角",
         blank="—", qty=1, unit="套",
         src="外購", lo=70, mid=110, hi=180,
         note="薄緣與前舌必須獨立保護"),
    dict(pn="WAVE-P03", name="PE袋、乾燥劑、銘牌、說明書", cat="包裝",
         spec="防塵袋＋10 g 乾燥劑×2＋金屬／木銘牌",
         blank="—", qty=1, unit="套",
         src="外購", lo=40, mid=70, hi=140,
         note="銘牌藏於盆底或後踵內側"),

    # 工裝攤提（小量 10 件攤）
    dict(pn="WAVE-T01", name="內盆陰模治具（MDF／PU）", cat="工裝攤提",
         spec="OP3 翻面吸附／定位",
         blank="1 套", qty=0.10, unit="套",
         src="自製", lo=800, mid=1400, hi=2200,
         note="治具總價約 8–20k，按 10 件攤；打樣則全計入首件"),
    dict(pn="WAVE-T02", name="組裝對位鋼治具", cat="工裝攤提",
         spec="A/B/C 止口＋銷孔基準",
         blank="1 套", qty=0.10, unit="套",
         src="自製", lo=400, mid=700, hi=1200,
         note="治具總價約 4–12k，按 10 件攤"),
]


def ntd(v):
    return round(v)


# ---------------------------------------------------------------------------
def style_header_row(ws, row, start, end, fill_color=C_HEAD, size=9):
    for col in range(start, end + 1):
        c = ws.cell(row, col)
        c.fill = fill(fill_color)
        c.font = font(size, True, "F4EFE6")
        c.alignment = align("center", "center")
        c.border = THIN


def apply_data_cell(c, h="left", size=9, bold=False, color=C_INK, bg=None):
    c.font = font(size, bold, color)
    c.alignment = align(h, "center")
    c.border = THIN
    if bg:
        c.fill = fill(bg)


def set_print(ws, title, landscape=True, fit_w=1, paper="A3"):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if paper == "A3" else ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = 1
    ws.page_setup.horizontalCentered = True
    ws.oddHeader.left.text = "&B WAVE Lounge Chair"
    ws.oddHeader.right.text = title
    ws.oddFooter.left.text = "WAVE-LC-001  BOM  Rev.A"
    ws.oddFooter.center.text = "概念工程 BOM，不可直接當採購單（未凍結 3D）"
    ws.oddFooter.right.text = "第 &P / &N 頁"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.6, header=0.3, footer=0.3)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:5"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110


def paint_banner(ws, last_col, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    a = ws["A1"]
    a.value = "WAVE  LOUNGE  CHAIR    工程用料清單  EBOM"
    a.font = font(16, True, "F4EFE6")
    a.fill = fill(C_HEAD)
    a.alignment = align("left", "center", False)
    b = ws["A2"]
    b.value = subtitle
    b.font = font(9, False, "F4EFE6")
    b.fill = fill(C_HEAD2)
    b.alignment = align("left", "center", False)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = fill(C_HEAD)
        ws.cell(2, col).fill = fill(C_HEAD2)


def info_bar(ws, row, last_col, items):
    """items: list of (label, value) spanning last_col cells."""
    # pack into last_col cells as pairs
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 18
    n = len(items)
    span = max(1, last_col // n)
    for i, (lab, val) in enumerate(items):
        c0 = 1 + i * span
        c1 = last_col if i == n - 1 else c0 + span - 1
        if c0 < c1:
            ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
            ws.merge_cells(start_row=row + 1, start_column=c0, end_row=row + 1, end_column=c1)
        cl = ws.cell(row, c0, lab)
        cv = ws.cell(row + 1, c0, val)
        cl.font = font(7, False, C_MUTED)
        cv.font = font(9, True, C_INK)
        cl.alignment = align("center", "center", False)
        cv.alignment = align("center", "center", False)
        cl.fill = fill(C_CREAM)
        cv.fill = fill(C_PAPER)
        for c in range(c0, c1 + 1):
            ws.cell(row, c).fill = fill(C_CREAM)
            ws.cell(row + 1, c).fill = fill(C_PAPER)
            ws.cell(row, c).border = THIN
            ws.cell(row + 1, c).border = THIN
            ws.cell(row, c).font = font(7, False, C_MUTED)
            ws.cell(row + 1, c).font = font(9, True, C_INK)
            ws.cell(row, c).alignment = align("center", "center", False)
            ws.cell(row + 1, c).alignment = align("center", "center", False)


def build():
    wb = Workbook()

    # ======================================================================
    # Sheet 0 — 封面
    # ======================================================================
    ws0 = wb.active
    ws0.title = "00_封面"
    ws0.sheet_properties.tabColor = C_HEAD
    for col, w in enumerate([28, 42, 22, 22, 22, 22], 1):
        ws0.column_dimensions[get_column_letter(col)].width = w
    paint_banner(ws0, 6, "圖號 WAVE-LC-001    版次 A    日期 2026-08-16    單位 NT$／mm    狀態：概念凍結前")
    ws0.row_dimensions[1].height = 28
    ws0.merge_cells("A4:F4")
    ws0["A4"] = "產品資訊"
    ws0["A4"].font = font(12, True, C_HEAD)
    meta = [
        ("產品名稱", "WAVE Lounge Chair 浪形實木休閒椅"),
        ("產品定位", "高階訂製／精品酒店／展場／客廳主視覺"),
        ("成品尺寸", "W 900  ×  D 960  ×  H 820 mm（座高含墊 400）"),
        ("主結構", "左殼 A ＋ 右殼 B ＋ 座盆 C，層積實木 5 軸加工"),
        ("主材", "歐洲／美國白橡 KD，含水率 8–10%（可改白蠟／山毛櫸／胡桃）"),
        ("表面", "硬質油／硬蠟油，開放顯紋；薄緣不做厚 PU"),
        ("軟包", "冷熟成型泡棉 60–90 mm ＋ 頭層牛皮或超纖"),
        ("BOM 基準", "成品 1 椅；成本以小量 10 件／件估算（工裝按 10 件攤）"),
        ("對應圖面", "WAVE-LC-001 概念三視圖 Rev.A"),
        ("編製", "材料顧問 ＋ 生產優化工程師"),
        ("注意", "3D 未凍結前，毛坯尺寸為估列，採購請加 15–20% 備料"),
    ]
    headers = ["項目", "內容"]
    ws0["A6"] = "項目"
    ws0["B6"] = "內容"
    ws0.merge_cells("B6:F6")
    style_header_row(ws0, 6, 1, 6)
    for i, (k, v) in enumerate(meta):
        r = 7 + i
        ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws0.cell(r, 1, k)
        ws0.cell(r, 2, v)
        bg = C_ROW_ALT if i % 2 else C_WHITE
        apply_data_cell(ws0.cell(r, 1), "left", 9, True, C_ACCENT, bg)
        for c in range(2, 7):
            apply_data_cell(ws0.cell(r, c), "left", 9, False, C_INK, bg)
        ws0.row_dimensions[r].height = 20

    ws0.merge_cells("A19:F19")
    ws0["A19"] = "分件示意"
    ws0["A19"].font = font(12, True, C_HEAD)
    parts = [
        ("料號", "名稱", "角色", "毛坯估（mm）", "淨材積估", "來源"),
        ("WAVE-A01", "左殼", "主承重＋造型", "500 × 900 × 220", "0.028–0.036 m³", "自製 CNC"),
        ("WAVE-B01", "右殼", "主承重＋造型", "500 × 900 × 220", "0.028–0.036 m³", "自製 CNC"),
        ("WAVE-C01", "座盆／脊板", "坐壓傳遞、藏縫", "560 × 620 × 80", "0.012–0.018 m³", "自製 CNC"),
        ("WAVE-U01/U02", "軟包總成", "人體工學", "依盆／背陰模", "—", "外購"),
        ("WAVE-H01", "調整腳墊 ×4", "接地微調", "Ø20–25", "—", "外購"),
    ]
    for i, row in enumerate(parts):
        r = 20 + i
        for c, val in enumerate(row, 1):
            cell = ws0.cell(r, c, val)
            if i == 0:
                apply_data_cell(cell, "center", 8, True, "F4EFE6", C_HEAD)
            else:
                bg = C_ROW_ALT if i % 2 else C_WHITE
                apply_data_cell(cell, "center" if c != 2 else "left", 9, False, C_INK, bg)
        ws0.row_dimensions[r].height = 20

    ws0.merge_cells("A27:F28")
    ws0["A27"] = (
        "本 BOM 為工程概念清單（EBOM），對應尚未凍結之 3D。"
        "木材毛坯、軟包開料、治具攤提將在分件模型確認後改為正式 MBOM／採購 BOM。"
        "單價為 2026 台灣市場粗估，實際以當期報價為準。"
    )
    ws0["A27"].font = font(8, False, C_MUTED)
    ws0["A27"].alignment = align("left", "center", True)
    ws0["A27"].fill = fill(C_CREAM)
    set_print(ws0, "封面", landscape=True, paper="A3")
    ws0.row_dimensions[3].height = 8

    # ======================================================================
    # Sheet 1 — BOM 總表
    # ======================================================================
    ws = wb.create_sheet("01_BOM總表")
    ws.sheet_properties.tabColor = C_ACCENT
    cols = [
        ("A", 6, "項次"),
        ("B", 14, "料號"),
        ("C", 36, "名稱"),
        ("D", 12, "類別"),
        ("E", 42, "規格／材質"),
        ("F", 22, "毛坯／用量"),
        ("G", 8, "數量"),
        ("H", 8, "單位"),
        ("I", 10, "來源"),
        ("J", 12, "單價低"),
        ("K", 12, "單價中"),
        ("L", 12, "單價高"),
        ("M", 12, "金額（中）"),
        ("N", 40, "工藝／備註"),
    ]
    last = len(cols)
    paint_banner(ws, last, "基準：成品 1 椅　　成本：小量 10 件攤工裝　　幣別：新台幣 NT$　　圖號 WAVE-LC-001 Rev.A")
    info_bar(ws, 3, last, [
        ("產品", "WAVE Lounge Chair"),
        ("成品尺寸", "900 × 960 × 820 mm"),
        ("主材", "白橡層積實木"),
        ("BOM 版次", "A  /  2026-08-16"),
        ("數量基準", "1 椅"),
        ("狀態", "概念／未凍結"),
        ("編製", "材料顧問＋生產優化"),
    ])

    head_row = 5
    for i, (_, w, title) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(head_row, i, title)
    style_header_row(ws, head_row, 1, last, C_HEAD, 8)
    ws.row_dimensions[head_row].height = 20
    ws.auto_filter.ref = f"A{head_row}:N{head_row + len(BOM)}"
    ws.freeze_panes = "A6"

    money_fmt = '#,##0'
    qty_fmt = '0.##'

    for i, item in enumerate(BOM, 1):
        r = head_row + i
        amount = item["qty"] * item["mid"] if item["cat"] != "工裝攤提" else item["mid"]
        # 工裝：mid 已是攤到單椅的金額；qty=0.10 套
        if item["cat"] == "工裝攤提":
            amount = item["mid"]  # already per-chair amortized
            # but mid is per-chair amortized already in the data
        else:
            # lo/mid/hi are already LINE totals for the qty listed (not unit price)
            # Wait - I defined lo/mid/hi as line cost for the listed qty, not unit price.
            # Looking at my data: qty=18 dowels, lo=80 mid=150 — that's LINE total.
            # qty=0.80 kg glue, mid=200 — LINE total.
            # So "單價" should be line/qty, 金額 = mid
            amount = item["mid"]
        unit_mid = item["mid"] / item["qty"] if item["qty"] else item["mid"]
        unit_lo = item["lo"] / item["qty"] if item["qty"] else item["lo"]
        unit_hi = item["hi"] / item["qty"] if item["qty"] else item["hi"]

        bg = C_CAT.get(item["cat"], C_WHITE)
        if i % 2 == 0:
            # slightly darker mix - just use cat color
            pass
        values = [
            i,
            item["pn"],
            item["name"],
            item["cat"],
            item["spec"],
            item["blank"],
            item["qty"],
            item["unit"],
            item["src"],
            unit_lo,
            unit_mid,
            unit_hi,
            amount,
            item["note"],
        ]
        hs = ["center", "center", "left", "center", "left", "left",
              "center", "center", "center", "right", "right", "right", "right", "left"]
        for c, (val, h) in enumerate(zip(values, hs), 1):
            cell = ws.cell(r, c, val)
            apply_data_cell(cell, h, 8, c in (2, 13), C_INK, bg)
        ws.cell(r, 7).number_format = qty_fmt
        for c in (10, 11, 12, 13):
            ws.cell(r, c).number_format = money_fmt
        ws.row_dimensions[r].height = 32

    # totals
    tr = head_row + len(BOM) + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=12)
    ws.cell(tr, 1, "單椅材料＋工裝攤提小計（中位估，不含人工／管銷／利潤）")
    ws.cell(tr, 1).font = font(9, True, "F4EFE6")
    ws.cell(tr, 1).fill = fill(C_HEAD)
    ws.cell(tr, 1).alignment = align("right", "center")
    for c in range(1, 13):
        ws.cell(tr, c).fill = fill(C_HEAD)
        ws.cell(tr, c).font = font(9, True, "F4EFE6")
        ws.cell(tr, c).border = THIN
    last_data = head_row + len(BOM)
    ws.cell(tr, 13, f"=SUM(M{head_row+1}:M{last_data})")
    ws.cell(tr, 13).font = font(10, True, "F4EFE6")
    ws.cell(tr, 13).fill = fill(C_HEAD)
    ws.cell(tr, 13).number_format = '"NT$"#,##0'
    ws.cell(tr, 13).alignment = align("right", "center")
    ws.cell(tr, 13).border = THIN
    ws.cell(tr, 14).fill = fill(C_HEAD)
    ws.cell(tr, 14).border = THIN
    ws.row_dimensions[tr].height = 22

    # range row
    tr2 = tr + 1
    lo_sum = sum(x["lo"] for x in BOM)
    hi_sum = sum(x["hi"] for x in BOM)
    ws.merge_cells(start_row=tr2, start_column=1, end_row=tr2, end_column=12)
    ws.cell(tr2, 1, f"區間估（低–高）：NT$ {lo_sum:,.0f}  –  NT$ {hi_sum:,.0f}    ※ 胡桃主材約再加 15–30%")
    ws.cell(tr2, 1).font = font(8, False, C_INK)
    ws.cell(tr2, 1).fill = fill(C_OAK)
    ws.cell(tr2, 1).alignment = align("right", "center")
    for c in range(1, last + 1):
        ws.cell(tr2, c).fill = fill(C_OAK)
        ws.cell(tr2, c).border = THIN
    ws.row_dimensions[tr2].height = 18

    note_r = tr2 + 2
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r + 2, end_column=last)
    ws.cell(note_r, 1, (
        "欄位說明：單價低／中／高為「該列數量」拆成單件單價；金額（中）＝該列中位總價。"
        "工裝攤提已按 10 件計入單椅。人工（CNC／砂光／組裝／塗裝）不在本表，見「03_成本彙總」。"
        "橡木單價依 KD FAS 約 NT$45,000–75,000／m³ 回推。軟包皮革價差最大，確認皮料後需重算 U03。"
    ))
    ws.cell(note_r, 1).font = font(8, False, C_MUTED)
    ws.cell(note_r, 1).alignment = align("left", "top", True)
    ws.row_dimensions[note_r].height = 18
    set_print(ws, "BOM 總表", landscape=True, paper="A3")
    ws.page_setup.fitToHeight = 1

    # ======================================================================
    # Sheet 2 — 木材下料
    # ======================================================================
    ws2 = wb.create_sheet("02_木材下料")
    ws2.sheet_properties.tabColor = "8B5A2B"
    paint_banner(ws2, 11, "層積實木毛坯　　含水率 8–10%　　單層 22–28 mm　　順紋沿緞帶　　白橡優先")
    info_bar(ws2, 3, 11, [
        ("樹種基準", "歐洲／美國白橡"),
        ("替代 1", "美國白蠟（較省）"),
        ("替代 2", "歐洲山毛櫸"),
        ("精品升級", "黑胡桃 ＋15–30%"),
        ("購料係數", "1.15–1.20（小量）"),
        ("窯乾", "KD，入廠養料 ≥7 天"),
    ])
    wood_heads = ["項次", "料號", "名稱", "毛坯 L", "毛坯 W", "毛坯 T",
                  "件數", "單件材積 m³", "小計 m³", "購料 m³（×1.18）", "備註"]
    for i, h in enumerate(wood_heads, 1):
        ws2.cell(5, i, h)
    style_header_row(ws2, 5, 1, 11, C_HEAD, 8)
    widths2 = [6, 14, 34, 12, 12, 12, 8, 14, 12, 16, 36]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wood_rows = [
        (1, "WAVE-A01", "左殼毛坯", 900, 500, 220, 1, "順紋沿緞帶；成對選料"),
        (2, "WAVE-B01", "右殼毛坯", 900, 500, 220, 1, "與 A01 同批、同色號"),
        (3, "WAVE-C01", "座盆毛坯", 620, 560, 80, 1, "盆底等效厚 ≥32"),
        (4, "WAVE-W09", "備料／補片", 800, 200, 28, 4, "薄緣補片、色差替換"),
    ]
    for i, (n, pn, name, L, W, T, qty, note) in enumerate(wood_rows):
        r = 6 + i
        vol = L * W * T / 1e9
        vals = [n, pn, name, L, W, T, qty, vol, f"=G{r}*H{r}", f"=I{r}*1.18", note]
        bg = C_ROW_ALT if i % 2 else C_WHITE
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            h = "center" if c not in (3, 11) else "left"
            apply_data_cell(cell, h, 9, c == 2, C_INK, bg)
        ws2.cell(r, 8).number_format = "0.000"
        ws2.cell(r, 9).number_format = "0.000"
        ws2.cell(r, 10).number_format = "0.000"
        ws2.row_dimensions[r].height = 22

    r = 10
    ws2.merge_cells("A10:H10")
    ws2["A10"] = "木材小計"
    ws2["A10"].font = font(9, True, "F4EFE6")
    for c in range(1, 9):
        ws2.cell(10, c).fill = fill(C_HEAD)
        ws2.cell(10, c).font = font(9, True, "F4EFE6")
        ws2.cell(10, c).border = THIN
        ws2.cell(10, c).alignment = align("right", "center")
    ws2["I10"] = "=SUM(I6:I9)"
    ws2["J10"] = "=SUM(J6:J9)"
    ws2["I10"].number_format = "0.000"
    ws2["J10"].number_format = "0.000"
    ws2["I10"].font = font(10, True, "F4EFE6")
    ws2["J10"].font = font(10, True, "F4EFE6")
    ws2["I10"].fill = fill(C_HEAD)
    ws2["J10"].fill = fill(C_HEAD)
    ws2["K10"].fill = fill(C_HEAD)
    for c in range(9, 12):
        ws2.cell(10, c).border = THIN
        ws2.cell(10, c).alignment = align("center", "center")

    # material cost by species
    ws2.merge_cells("A12:K12")
    ws2["A12"] = "樹種材料費回推（依購料材積 J10 × 市場單價）"
    ws2["A12"].font = font(11, True, C_HEAD)

    sp_heads = ["樹種", "單價低 NT$/m³", "單價中", "單價高", "材料費低", "材料費中", "材料費高", "適用", "", "", ""]
    for i, h in enumerate(["樹種", "單價低 NT$/m³", "單價中", "單價高",
                           "材料費低", "材料費中", "材料費高", "建議用途"], 1):
        ws2.cell(13, i, h)
    style_header_row(ws2, 13, 1, 8, C_HEAD2, 8)

    species = [
        ("美國白蠟 Ash", 32000, 42000, 52000, "量產／淺色、彈性好"),
        ("歐洲山毛櫸", 28000, 36000, 45000, "均質好雕、移動較大"),
        ("白橡 White Oak（基準）", 45000, 58000, 75000, "本專案主選，接近參考圖"),
        ("歐洲橡 Quartersawn", 55000, 70000, 90000, "腹板更穩、價較高"),
        ("黑胡桃 Walnut", 85000, 105000, 130000, "精品溢價，陰影較重"),
    ]
    for i, (name, lo, mid, hi, use) in enumerate(species):
        r = 14 + i
        ws2.cell(r, 1, name)
        ws2.cell(r, 2, lo)
        ws2.cell(r, 3, mid)
        ws2.cell(r, 4, hi)
        ws2.cell(r, 5, f"=$J$10*B{r}")
        ws2.cell(r, 6, f"=$J$10*C{r}")
        ws2.cell(r, 7, f"=$J$10*D{r}")
        ws2.cell(r, 8, use)
        bg = "E8D4B5" if "基準" in name else (C_ROW_ALT if i % 2 else C_WHITE)
        for c in range(1, 9):
            apply_data_cell(ws2.cell(r, c), "center" if c != 8 and c != 1 else "left", 9, "基準" in name, C_INK, bg)
        for c in range(2, 8):
            ws2.cell(r, c).number_format = "#,##0"
        ws2.row_dimensions[r].height = 20

    ws2.merge_cells("A20:K22")
    ws2["A20"] = (
        "下料規則：\n"
        "1. 單片拼板寬 ≤120 mm；層積單層 22–28 mm；層間交錯年輪。\n"
        "2. 禁止獨幅厚材直接上 5 軸。毛坯六面基準後做 2× Ø8 H7 工藝孔。\n"
        "3. 利用率目標 25–40%（分件層積）。整塊掏空不採用。\n"
        "4. 雕刻後恆濕 48–72 h 再精修。端紋與接地面先封孔。"
    )
    ws2["A20"].font = font(8, False, C_INK)
    ws2["A20"].alignment = align("left", "top", True)
    ws2["A20"].fill = fill(C_CREAM)
    ws2.row_dimensions[20].height = 20
    ws2.row_dimensions[21].height = 20
    ws2.row_dimensions[22].height = 20
    ws2.freeze_panes = "A6"
    set_print(ws2, "木材下料", landscape=True, paper="A3")

    # ======================================================================
    # Sheet 3 — 成本彙總
    # ======================================================================
    ws3 = wb.create_sheet("03_成本彙總")
    ws3.sheet_properties.tabColor = C_GREEN
    paint_banner(ws3, 8, "小量 10 件／件　　打樣 1 件另列　　不含品牌溢價、設計費、運費")
    info_bar(ws3, 3, 8, [
        ("幣別", "新台幣 NT$"),
        ("木材基準", "白橡中位"),
        ("軟包基準", "中配（超纖偏頭層中位）"),
        ("工裝", "10 件攤"),
    ])

    # category rollup
    cat_order = ["結構木材", "接合件", "五金", "膠合", "塗裝", "軟包", "包裝", "工裝攤提"]
    cat_sum = {c: [0, 0, 0] for c in cat_order}
    for it in BOM:
        cat_sum[it["cat"]][0] += it["lo"]
        cat_sum[it["cat"]][1] += it["mid"]
        cat_sum[it["cat"]][2] += it["hi"]

    ws3["A5"] = "一、單椅直接材料（含工裝攤提）"
    ws3["A5"].font = font(11, True, C_HEAD)
    ws3.merge_cells("A5:E5")
    for i, h in enumerate(["類別", "低", "中", "高", "佔中位 %"], 1):
        ws3.cell(6, i, h)
    style_header_row(ws3, 6, 1, 5, C_HEAD, 9)
    mid_total = sum(v[1] for v in cat_sum.values())
    for i, cat in enumerate(cat_order):
        r = 7 + i
        lo, mid, hi = cat_sum[cat]
        ws3.cell(r, 1, cat)
        ws3.cell(r, 2, lo)
        ws3.cell(r, 3, mid)
        ws3.cell(r, 4, hi)
        ws3.cell(r, 5, mid / mid_total if mid_total else 0)
        bg = C_CAT[cat]
        for c in range(1, 6):
            apply_data_cell(ws3.cell(r, c), "center" if c > 1 else "left", 9, False, C_INK, bg)
        for c in range(2, 5):
            ws3.cell(r, c).number_format = "#,##0"
        ws3.cell(r, 5).number_format = "0.0%"
        ws3.row_dimensions[r].height = 20

    r = 15
    ws3.cell(r, 1, "材料小計")
    ws3.cell(r, 2, f"=SUM(B7:B14)")
    ws3.cell(r, 3, f"=SUM(C7:C14)")
    ws3.cell(r, 4, f"=SUM(D7:D14)")
    ws3.cell(r, 5, 1)
    for c in range(1, 6):
        ws3.cell(r, c).font = font(9, True, "F4EFE6")
        ws3.cell(r, c).fill = fill(C_HEAD)
        ws3.cell(r, c).alignment = align("center" if c > 1 else "left", "center")
        ws3.cell(r, c).border = THIN
    for c in range(2, 5):
        ws3.cell(r, c).number_format = '"NT$"#,##0'
    ws3.cell(r, 5).number_format = "0%"

    # labor
    ws3["A17"] = "二、單椅直接人工（台製熟手粗估，不含管銷）"
    ws3["A17"].font = font(11, True, C_HEAD)
    ws3.merge_cells("A17:E17")
    for i, h in enumerate(["工段", "打樣工時 h", "小量工時 h", "小量人工 NT$（中）", "備註"], 1):
        ws3.cell(18, i, h)
    style_header_row(ws3, 18, 1, 5, C_HEAD2, 8)
    labor = [
        ("選料拼板層積", 5.0, 3.0, 1800, "含養膠"),
        ("CNC 機時（A+B+C）", 10.0, 6.0, 7200, "含換刀換治具；機時費另見機折舊"),
        ("組裝校正", 2.5, 1.2, 720, "對位治具"),
        ("砂光", 16.0, 10.0, 6000, "工時黑洞，精修刀路可降"),
        ("塗裝 2–3 遍", 3.5, 3.0, 1500, "乾燥等待另計產線占用"),
        ("軟包配合／總裝", 2.0, 1.2, 720, "若軟包全外包則降至 0.5 h"),
    ]
    # labor mid = small-batch hours * ~NT$600/h blended? Let me use explicit amounts
    # Recalculate more consistently: 小量人工 using ~NT$550-650/h for shop labor
    # CNC machine rate separate: ~NT$800-1200/h machine
    # I'll split: 人工工資 vs CNC 機時費
    labor = [
        ("選料拼板層積（人工）", 5.0, 3.0, 3.0 * 600, "含養膠"),
        ("CNC 機時費（折舊＋電＋刀）", 10.0, 6.0, 6.0 * 1200, "非工資；機台小時費"),
        ("CNC 操作人工", 10.0, 6.0, 6.0 * 450, "一人看機"),
        ("組裝校正", 2.5, 1.2, 1.2 * 600, "對位治具"),
        ("砂光", 16.0, 10.0, 10.0 * 600, "最大成本風險"),
        ("塗裝", 3.5, 3.0, 3.0 * 550, "不含乾燥等待工資"),
        ("軟包配合／總裝", 2.0, 1.2, 1.2 * 600, "軟包本體已在材料"),
    ]
    for i, (name, proto_h, batch_h, ntd_mid, note) in enumerate(labor):
        r = 19 + i
        ws3.cell(r, 1, name)
        ws3.cell(r, 2, proto_h)
        ws3.cell(r, 3, batch_h)
        ws3.cell(r, 4, ntd_mid)
        ws3.cell(r, 5, note)
        bg = C_ROW_ALT if i % 2 else C_WHITE
        for c in range(1, 6):
            apply_data_cell(ws3.cell(r, c), "center" if c in (2, 3, 4) else "left", 9, False, C_INK, bg)
        ws3.cell(r, 2).number_format = "0.0"
        ws3.cell(r, 3).number_format = "0.0"
        ws3.cell(r, 4).number_format = "#,##0"
        ws3.row_dimensions[r].height = 20
    r = 26
    ws3.cell(r, 1, "人工＋機時小計（小量）")
    ws3.cell(r, 2, "=SUM(B19:B25)")
    ws3.cell(r, 3, "=SUM(C19:C25)")
    ws3.cell(r, 4, "=SUM(D19:D25)")
    for c in range(1, 6):
        ws3.cell(r, c).font = font(9, True, "F4EFE6")
        ws3.cell(r, c).fill = fill(C_HEAD2)
        ws3.cell(r, c).border = THIN
        ws3.cell(r, c).alignment = align("center" if c > 1 else "left", "center")
    ws3.cell(r, 2).number_format = "0.0"
    ws3.cell(r, 3).number_format = "0.0"
    ws3.cell(r, 4).number_format = '"NT$"#,##0'

    # grand
    ws3["A28"] = "三、單椅製造成本滾算"
    ws3["A28"].font = font(11, True, C_HEAD)
    ws3.merge_cells("A28:E28")
    for i, h in enumerate(["項目", "打樣 1 件", "小量 10 件／件", "備註"], 1):
        ws3.cell(29, i, h)
    style_header_row(ws3, 29, 1, 4, C_HEAD, 9)
    # Prototype: materials without amortization * full jig + higher labor
    mat_mid = mid_total
    jig_full_extra = (1400 + 700) * 9  # remaining 9/10 of jig if first piece bears all
    # cleaner:
    roll = [
        ("直接材料（含 10 件攤工裝）", "C15", "C15", "打樣若工裝全計入，另加約 18–28k"),
        ("直接人工＋CNC 機時（小量）", "D26*1.55", "D26", "打樣工時約為小量 1.5–1.8 倍"),
        ("不良／備料風險 8%", "(C15+D26*1.55)*0.08", "(C15+D26)*0.08", "薄緣崩缺、色差、對接重工"),
        ("製造成本小計", None, None, "不含管銷、設計、運費"),
        ("管銷 18%", None, None, "場租、管理、品管"),
        ("建議出廠底價（再加利潤前）", None, None, "利潤建議再加 20–35%"),
    ]
    # Use numeric computed values for clarity rather than fragile cross-formula
    proto_labor = 5*600 + 10*1200 + 10*450 + 2.5*600 + 16*600 + 3.5*550 + 2*600
    batch_labor = sum(x[3] for x in labor)
    proto_mat = mat_mid + 18900  # full jig on first piece roughly (14k+7k - already 2.1k in mat)
    # jig in BOM mid = 1400+700=2100 (1/10). Full jig = 21000. Extra on proto = 18900
    proto_risk = (proto_mat + proto_labor) * 0.10
    batch_risk = (mat_mid + batch_labor) * 0.08
    proto_mfg = proto_mat + proto_labor + proto_risk
    batch_mfg = mat_mid + batch_labor + batch_risk
    proto_oh = proto_mfg * 0.18
    batch_oh = batch_mfg * 0.18
    proto_floor = proto_mfg + proto_oh
    batch_floor = batch_mfg + batch_oh

    roll_vals = [
        ("直接材料", proto_mat, mat_mid, "打樣含治具全額；小量已攤 1/10"),
        ("直接人工＋CNC 機時", proto_labor, batch_labor, "打樣工時較高、CAM 另計一次性 40–80 h"),
        ("不良／備料風險", proto_risk, batch_risk, "打樣 10%／小量 8%"),
        ("製造成本小計", proto_mfg, batch_mfg, "不含管銷、設計、運費、CAM 建模"),
        ("管銷 18%", proto_oh, batch_oh, "場租、管理、品管"),
        ("建議出廠底價（未加利潤）", proto_floor, batch_floor, "利潤建議再加 20–35%"),
        ("建議出廠價（含約 25% 利潤）", proto_floor * 1.25, batch_floor * 1.25, "對客報價參考，非正式報價單"),
    ]
    for i, (name, proto, batch, note) in enumerate(roll_vals):
        r = 30 + i
        ws3.cell(r, 1, name)
        ws3.cell(r, 2, round(proto))
        ws3.cell(r, 3, round(batch))
        ws3.cell(r, 4, note)
        last_rows = i >= 3
        bold = i >= 3
        bg = C_HEAD if i == 6 else (C_OAK if last_rows else (C_ROW_ALT if i % 2 else C_WHITE))
        fg = "F4EFE6" if i == 6 else C_INK
        for c in range(1, 5):
            apply_data_cell(ws3.cell(r, c), "center" if c in (2, 3) else "left", 9, bold, fg, bg)
        ws3.cell(r, 2).number_format = '"NT$"#,##0'
        ws3.cell(r, 3).number_format = '"NT$"#,##0'
        ws3.row_dimensions[r].height = 22

    ws3.merge_cells("A38:E40")
    ws3["A38"] = (
        "報價指引：白橡＋中配軟包，小量建議出廠 NT$70,000–130,000／椅；打樣 NT$120,000–200,000。"
        "胡桃或頭層高配皮另議。CAM／建模為一次性費用，建議報價時獨立成項，勿攤進第一件後消失。"
        "本表可作內部成本，不可直接當客戶報價單。"
    )
    ws3["A38"].font = font(8, False, C_MUTED)
    ws3["A38"].alignment = align("left", "top", True)
    ws3["A38"].fill = fill(C_CREAM)

    for i, w in enumerate([36, 16, 16, 18, 42, 14, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A7"
    set_print(ws3, "成本彙總", landscape=True, paper="A3")

    # chart data already in sheet — add bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "單椅材料成本結構（中位 NT$）"
    chart.y_axis.title = "NT$"
    chart.x_axis.title = None
    data = Reference(ws3, min_col=3, min_row=6, max_row=14)
    cats = Reference(ws3, min_col=1, min_row=7, max_row=14)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.legend = None
    chart.width = 18
    chart.height = 8
    ws3.add_chart(chart, "A42")

    # ======================================================================
    # Sheet 4 — 採購清單（外購 only）
    # ======================================================================
    ws4 = wb.create_sheet("04_外購清單")
    ws4.sheet_properties.tabColor = "4A6B8A"
    paint_banner(ws4, 9, "僅列外購項　　自製件見 01／02　　可直接轉請購")
    info_bar(ws4, 3, 9, [
        ("用途", "請購／比價"),
        ("數量基準", "1 椅；下單請乘批量"),
        ("建議批量", "10 椅＋備料 15%"),
        ("幣別", "NT$"),
    ])
    p_heads = ["項次", "料號", "名稱", "規格", "每椅用量", "單位", "10 椅＋15%", "預估單價中", "10 椅金額中"]
    for i, h in enumerate(p_heads, 1):
        ws4.cell(5, i, h)
    style_header_row(ws4, 5, 1, 9, C_HEAD, 8)
    widths4 = [6, 14, 32, 44, 12, 8, 14, 14, 14]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    buy = [it for it in BOM if it["src"] == "外購" and it["cat"] != "工裝攤提"]
    for i, it in enumerate(buy):
        r = 6 + i
        unit_mid = it["mid"] / it["qty"] if it["qty"] else it["mid"]
        batch_qty = it["qty"] * 10 * 1.15
        vals = [i + 1, it["pn"], it["name"], it["spec"], it["qty"], it["unit"],
                batch_qty, unit_mid, batch_qty * unit_mid]
        bg = C_ROW_ALT if i % 2 else C_WHITE
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(r, c, v)
            apply_data_cell(cell, "center" if c not in (3, 4) else "left", 8, c == 2, C_INK, bg)
        ws4.cell(r, 5).number_format = "0.##"
        ws4.cell(r, 7).number_format = "0.##"
        ws4.cell(r, 8).number_format = "#,##0.00"
        ws4.cell(r, 9).number_format = "#,##0"
        ws4.row_dimensions[r].height = 28

    end = 5 + len(buy)
    tot = end + 1
    ws4.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=8)
    ws4.cell(tot, 1, "10 椅外購小計（中位，含 15% 備料）")
    ws4.cell(tot, 1).font = font(9, True, "F4EFE6")
    ws4.cell(tot, 1).fill = fill(C_HEAD)
    ws4.cell(tot, 1).alignment = align("right", "center")
    for c in range(1, 9):
        ws4.cell(tot, c).fill = fill(C_HEAD)
        ws4.cell(tot, c).border = THIN
    ws4.cell(tot, 9, f"=SUM(I6:I{end})")
    ws4.cell(tot, 9).font = font(10, True, "F4EFE6")
    ws4.cell(tot, 9).fill = fill(C_HEAD)
    ws4.cell(tot, 9).number_format = '"NT$"#,##0'
    ws4.cell(tot, 9).border = THIN
    ws4.cell(tot, 9).alignment = align("right", "center")
    ws4.freeze_panes = "A6"
    ws4.auto_filter.ref = f"A5:I{end}"
    set_print(ws4, "外購清單", landscape=True, paper="A3")

    # ======================================================================
    # Sheet 5 — 修訂紀錄
    # ======================================================================
    ws5 = wb.create_sheet("05_修訂紀錄")
    ws5.sheet_properties.tabColor = "6B6560"
    paint_banner(ws5, 5, "文件管制")
    for i, w in enumerate([10, 16, 16, 18, 60], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(["版次", "日期", "編製", "核准", "說明"], 1):
        ws5.cell(4, i, h)
    style_header_row(ws5, 4, 1, 5)
    ws5.cell(5, 1, "A")
    ws5.cell(5, 2, "2026-08-16")
    ws5.cell(5, 3, "材料顧問／生產優化")
    ws5.cell(5, 4, "待主設計師確認")
    ws5.cell(5, 5, "初版概念 EBOM。依 WAVE-LC-001 Rev.A 控制尺寸與分件 A/B/C 建立。3D 凍結後升版 B。")
    for c in range(1, 6):
        apply_data_cell(ws5.cell(5, c), "center" if c < 5 else "left", 9, False, C_INK, C_WHITE)
    ws5.row_dimensions[5].height = 36
    ws5.merge_cells("A7:E10")
    ws5["A7"] = (
        "升版條件（任一發生即出 Rev.B）：\n"
        "• 樹種或表面處理定案與本表不同\n"
        "• 分件由 A/B/C 改為其他切法\n"
        "• 軟包定案為一體或分件、皮／布確認\n"
        "• 3D 凍結、毛坯尺寸改為實量"
    )
    ws5["A7"].font = font(9, False, C_INK)
    ws5["A7"].alignment = align("left", "top", True)
    set_print(ws5, "修訂紀錄", landscape=False, paper="A4")

    # freeze / print area
    xlsx = OUT / "WAVE-LC-001_BOM_RevA.xlsx"
    wb.save(xlsx)

    # CSV of master BOM
    import csv
    csv_path = OUT / "WAVE-LC-001_BOM_RevA.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["項次", "料號", "名稱", "類別", "規格／材質", "毛坯／用量",
                    "數量", "單位", "來源", "金額低", "金額中", "金額高", "備註"])
        for i, it in enumerate(BOM, 1):
            w.writerow([i, it["pn"], it["name"], it["cat"], it["spec"], it["blank"],
                        it["qty"], it["unit"], it["src"], it["lo"], it["mid"], it["hi"], it["note"]])
        w.writerow([])
        w.writerow(["小計", "", "", "", "", "", "", "", "",
                    sum(x["lo"] for x in BOM), sum(x["mid"] for x in BOM),
                    sum(x["hi"] for x in BOM), "NT$ 單椅材料+工裝攤提"])

    return xlsx, csv_path, cat_sum, mid_total, proto_floor, batch_floor


if __name__ == "__main__":
    xlsx, csv_path, cat_sum, mid_total, proto_floor, batch_floor = build()
    print(xlsx)
    print(csv_path)
    print("MAT_MID", mid_total)
    print("PROTO_FLOOR", round(proto_floor))
    print("BATCH_FLOOR", round(batch_floor))
    for k, v in cat_sum.items():
        print(f"{k}\t{v[0]}\t{v[1]}\t{v[2]}")
